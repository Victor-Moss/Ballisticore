import openpyxl
from openpyxl import load_workbook

FILE_PATH = "C:/sites/BallistiCore/Excel_Sample/Firearms Permit.xlsm"

print("=" * 80)
print("LOADING WORKBOOK:", FILE_PATH)
print("=" * 80)

wb = load_workbook(FILE_PATH, keep_vba=True, data_only=True)

print(f"\nTotal sheets found: {len(wb.sheetnames)}")
print("\nSHEET NAMES:")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {name}")

print("\n" + "=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: '{sheet_name}'")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    print("-" * 80)

    rows_to_read = min(60, ws.max_row)

    # First pass: find all non-empty rows and collect column widths for display
    all_rows = []
    for row_idx in range(1, rows_to_read + 1):
        row_data = []
        has_value = False
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                has_value = True
            row_data.append(val)
        all_rows.append((row_idx, row_data, has_value))

    # Print each row
    empty_count = 0
    for row_idx, row_data, has_value in all_rows:
        if not has_value:
            empty_count += 1
            if empty_count <= 2:
                print(f"  Row {row_idx:3d}: [EMPTY]")
            elif empty_count == 3:
                print(f"  ...")
            continue
        empty_count = 0

        # Format the row
        parts = []
        for col_idx, val in enumerate(row_data, 1):
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                # Truncate long values
                str_val = str(val)
                if len(str_val) > 60:
                    str_val = str_val[:57] + "..."
                parts.append(f"{col_letter}={str_val!r}")

        if parts:
            print(f"  Row {row_idx:3d}: {' | '.join(parts)}")

    print(f"\n  [Sheet '{sheet_name}' summary: max_row={ws.max_row}, max_col={ws.max_column}]")

    # Print merged cells if any
    if ws.merged_cells.ranges:
        print(f"\n  MERGED CELLS:")
        for merged in list(ws.merged_cells.ranges)[:20]:
            print(f"    {merged}")

    # Print named ranges scoped to this sheet
    # (handled globally below)

print("\n" + "=" * 80)
print("NAMED RANGES / DEFINED NAMES:")
print("=" * 80)
try:
    for name, defn in wb.defined_names.items():
        print(f"  {name}: {defn.attr_text}")
except Exception as e:
    print(f"  (Could not read defined names: {e})")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
