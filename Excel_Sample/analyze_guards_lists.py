import openpyxl

FILE_PATH = "C:/sites/BallistiCore/Excel_Sample/Firearms Permit.xlsm"
wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True, data_only=True)

# ----- GUARDS sheet: read rows 3-30 (header + sample data) -----
print("=" * 80)
print("GUARDS SHEET - Full column headers (row 3) and data rows 4-30")
print("=" * 80)
ws = wb["Guards"]
for row_idx in range(3, 31):
    row_data = []
    has_value = False
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None and str(val).strip() not in ('', '.'):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}={val!r}")
            has_value = True
    if has_value:
        print(f"  Row {row_idx:3d}: {' | '.join(row_data)}")
    else:
        print(f"  Row {row_idx:3d}: [empty]")

# ----- LISTS sheet: read rows 1-10 (firearm structure) -----
print("\n" + "=" * 80)
print("LISTS SHEET - Firearm columns (rows 1-10)")
print("=" * 80)
ws = wb["Lists"]
for row_idx in range(1, 11):
    row_data = []
    has_value = False
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None and str(val).strip() not in ('', '.'):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}={val!r}")
            has_value = True
    if has_value:
        print(f"  Row {row_idx:3d}: {' | '.join(row_data)}")

# ----- SYSTEM ADMIN: show all data rows fully -----
print("\n" + "=" * 80)
print("SYSTEM ADMIN SHEET - Full user rows")
print("=" * 80)
ws = wb["System Admin"]
for row_idx in range(1, 10):
    row_data = []
    has_value = False
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None and str(val).strip() not in ('', '.'):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_data.append(f"{col_letter}={val!r}")
            has_value = True
    if has_value:
        print(f"  Row {row_idx:3d}: {' | '.join(row_data)}")
    else:
        print(f"  Row {row_idx:3d}: [empty]")

# ----- MAIN MENU: specific config cells -----
print("\n" + "=" * 80)
print("MAIN MENU - Config/Settings cells (rows 1-20)")
print("=" * 80)
ws = wb["Main Menu"]
for row_idx in range(1, 21):
    row_data = []
    has_value = False
    for col_idx in range(1, 40):  # limit columns for readability
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None and str(val).strip() not in ('', '.'):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            str_val = str(val)[:80]
            row_data.append(f"{col_letter}={str_val!r}")
            has_value = True
    if has_value:
        print(f"  Row {row_idx:3d}: {' | '.join(row_data)}")
    else:
        print(f"  Row {row_idx:3d}: [empty]")
