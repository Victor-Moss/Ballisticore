#!/usr/bin/env python3
"""
Analyze the Excel sample file (Firearms Permit.xlsm) in detail.
Extract worksheets, data, formulas, and VBA code.
"""

import os
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import pathlib

EXCEL_FILE = "Excel_Sample/Firearms Permit.xlsm"

def analyze_excel():
    print("=== EXCEL FILE ANALYSIS ===")
    print(f"File: {EXCEL_FILE}")

    # Load workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)
    print(f"Worksheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Worksheet: {sheet_name} ---")
        print(f"Dimensions: {sheet.dimensions}")

        # Get used range
        min_row, min_col, max_row, max_col = sheet.min_row, sheet.min_column, sheet.max_row, sheet.max_column
        print(f"Used range: {min_row}:{min_col} to {max_row}:{max_col}")

        # Sample data (first 10 rows)
        print("Sample data:")
        for row in range(min_row, min(min_row + 10, max_row + 1)):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    row_data.append(str(value)[:50])  # Truncate long values
                else:
                    row_data.append("")
            if any(row_data):  # Only print non-empty rows
                print(f"Row {row}: {row_data}")

    # Now extract VBA code from the .xlsm file (it's a zip archive)
    print("\n=== VBA CODE ANALYSIS ===")
    try:
        with zipfile.ZipFile(EXCEL_FILE, 'r') as zf:
            # Look for VBA files
            vba_files = [f for f in zf.namelist() if 'vba' in f.lower() or f.endswith('.bas') or f.endswith('.cls')]
            print(f"VBA files found: {vba_files}")

            for vba_file in vba_files:
                print(f"\n--- VBA File: {vba_file} ---")
                try:
                    with zf.open(vba_file) as f:
                        content = f.read().decode('utf-8', errors='ignore')
                        print(content[:2000])  # First 2000 chars
                        if len(content) > 2000:
                            print("... (truncated)")
                except Exception as e:
                    print(f"Error reading {vba_file}: {e}")

    except Exception as e:
        print(f"Error extracting VBA: {e}")

if __name__ == "__main__":
    analyze_excel()