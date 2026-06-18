"""Tests for the full data export (Excel + CSV bundle + PDF summary)."""
import uuid
import zipfile
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.services import exports as svc
from app.services.imports import SHEETS
from app.models.guard import Guard
from app.models.user import User
from app.models.firearm import Firearm
from app.models.register_history import RegisterHistory


def _seed(db):
    future = date.today() + timedelta(days=200)
    past = date.today() - timedelta(days=10)
    db.add(Guard(first_name="Alice", last_name="Active", psira_number="PS1",
                 saps_comp_shotgun="C7021766", saps_expiry_shotgun=future, permitted_shotgun=True))
    db.add(Guard(first_name="Bob", last_name="Lapsed", psira_number=None,
                 saps_comp_rifle="E2345678", saps_expiry_rifle=past, permitted_rifle=True))
    db.commit()


def test_excel_has_a_sheet_per_entity_and_guards_match_import_columns(db):
    _seed(db)
    datasets = svc.collect_datasets(db)
    wb = load_workbook(BytesIO(svc.build_excel(datasets)))

    expected = {"Guards", "Users", "Firearms", "Locations", "Ammunition Types",
                "Guard Firearm Permissions", "Guard CIT Routes", "Permits",
                "Register", "Register History"}
    assert expected.issubset(set(wb.sheetnames))

    # Guards columns are derived from the import template — they must match.
    import_headers = [h for (h, *_ ) in next(s for s in SHEETS if s["name"] == "Guards")["columns"]]
    guard_headers = [c.value for c in wb["Guards"][1]]
    assert guard_headers == import_headers

    # Two seeded guards present.
    assert wb["Guards"].max_row == 3  # header + 2


def test_user_export_excludes_password_hash(db):
    _seed(db)
    datasets = svc.collect_datasets(db)
    users_ds = next(d for d in datasets if d["name"] == "Users")
    assert "hashed_password" not in users_ds["headers"]


def _ds(datasets, name):
    return next(d for d in datasets if d["name"] == name)


def test_fk_companion_columns_in_register_history(db):
    guard = Guard(id=str(uuid.uuid4()), first_name="Carl", last_name="Carrier")
    fa = Firearm(id=str(uuid.uuid4()), serial_number="GLK-100234", make="Glock",
                 model="17", type="handgun")
    user = User(id=str(uuid.uuid4()), username="operator1",
                hashed_password="x", is_active=True)
    db.add_all([guard, fa, user])
    db.commit()
    db.add(RegisterHistory(id=str(uuid.uuid4()), guard_id=guard.id, firearm_id=fa.id,
                           action="ISSUED", actioned_by=user.id))
    db.commit()

    ds = _ds(svc.collect_datasets(db), "Register History")
    headers, row = ds["headers"], ds["rows"][0]
    rowmap = dict(zip(headers, row))

    # Companion column sits immediately after its FK column, ID kept intact.
    assert headers[headers.index("guard_id") + 1] == "guard_name"
    assert headers[headers.index("firearm_id") + 1] == "firearm_description"
    assert headers[headers.index("actioned_by") + 1] == "actioned_by_name"
    assert rowmap["guard_id"] == guard.id            # raw UUID preserved
    assert rowmap["guard_name"] == "Carl Carrier"
    assert rowmap["firearm_description"] == "GLK-100234 (Glock 17, handgun)"
    assert rowmap["actioned_by_name"] == "operator1"


def test_no_fk_entities_have_no_companion_columns(db):
    # Entities with no foreign keys must export exactly their own columns —
    # no extra companion columns added.
    from app.models.location import Location
    from app.models.ammunition_type import AmmunitionType
    _seed(db)
    datasets = svc.collect_datasets(db)
    for name, model in (("Users", User), ("Locations", Location), ("Ammunition Types", AmmunitionType)):
        expected = [c.name for c in model.__table__.columns if c.name not in svc._SECRET_COLUMNS]
        assert _ds(datasets, name)["headers"] == expected


def test_guards_sheet_unchanged_no_fk_uuids(db):
    # Guards columns come from the import template, which omits location_id, so
    # there are no opaque FK UUIDs to humanise — the sheet stays as-is.
    _seed(db)
    headers = _ds(svc.collect_datasets(db), "Guards")["headers"]
    import_headers = [h for (h, *_ ) in next(s for s in SHEETS if s["name"] == "Guards")["columns"]]
    assert headers == import_headers
    assert not any(h.endswith("_id") for h in headers)


def test_csv_zip_has_one_file_per_entity(db):
    _seed(db)
    datasets = svc.collect_datasets(db)
    z = zipfile.ZipFile(BytesIO(svc.build_csv_zip(datasets)))
    names = set(z.namelist())
    assert "guards.csv" in names
    assert "register_history.csv" in names
    assert len(names) == len(datasets)
    # Guards CSV has a header + the two seeded rows.
    rows = z.read("guards.csv").decode().strip().splitlines()
    assert len(rows) == 3


def test_pdf_summary_is_a_pdf(db):
    _seed(db)
    pdf = svc.build_pdf_summary(db, generated_by="tester", generated_at=datetime.now())
    assert pdf[:4] == b"%PDF"


def test_full_export_bundles_three_artefacts(db):
    _seed(db)
    data, filename = svc.build_full_export(db, generated_by="tester")
    assert filename.endswith(".zip")
    z = zipfile.ZipFile(BytesIO(data))
    names = z.namelist()
    assert any(n.endswith(".xlsx") for n in names)
    assert any(n.endswith(".zip") for n in names)   # the CSV bundle
    assert any(n.endswith(".pdf") for n in names)


def test_full_export_logs_audit_event(db, caplog):
    _seed(db)
    import logging
    with caplog.at_level(logging.INFO, logger="ballisticore.export"):
        svc.build_full_export(db, generated_by="auditor")
    assert any("auditor" in r.getMessage() for r in caplog.records)


def test_export_endpoint_requires_admin(client, db):
    from tests.conftest import make_user, auth_headers
    operator = make_user(db, username="op", is_admin=False)
    res = client.get("/api/export/all", headers=auth_headers(operator.id))
    assert res.status_code == 403


def test_export_endpoint_allows_admin(client, db):
    from tests.conftest import make_user, auth_headers
    admin = make_user(db, username="boss", is_admin=True)
    res = client.get("/api/export/all", headers=auth_headers(admin.id))
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/zip"
    assert res.content[:2] == b"PK"  # ZIP magic
