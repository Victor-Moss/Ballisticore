"""Tests for SAPS competency handling in the Excel bulk import."""
import base64
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services import imports as svc
from app.models.guard import Guard


# Header order must match the Guards sheet template exactly.
GUARD_HEADERS = [h for (h, *_ ) in next(s for s in svc.SHEETS if s["name"] == "Guards")["columns"]]


def _make_guards_book(rows: list[dict]) -> bytes:
    """Build an .xlsx with a Guards sheet from a list of {header: value} dicts."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Guards")
    ws.append(GUARD_HEADERS)
    for r in rows:
        ws.append([r.get(h, None) for h in GUARD_HEADERS])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _guard(db, last_name):
    return db.query(Guard).filter(Guard.last_name == last_name).first()


def test_valid_pair_auto_ticks_permission(db):
    future = date.today() + timedelta(days=365)
    data = _make_guards_book([{
        "First Name *": "Alice", "Last Name *": "Auto",
        "Shotgun Competency Number": "c7021766",  # lowercase -> normalised
        "Shotgun Competency Expiry": future,
    }])
    res = svc.import_workbook(db, data)
    assert res["imported"] == 1
    assert res["failed"] == 0
    g = _guard(db, "Auto")
    assert g.permitted_shotgun is True
    assert g.saps_comp_shotgun == "C7021766"      # uppercased
    assert g.saps_expiry_shotgun == future
    assert g.permitted_rifle is False             # untouched weapons stay off


def test_expired_pair_imports_ticks_and_is_flagged(db):
    past = date.today() - timedelta(days=10)
    data = _make_guards_book([{
        "First Name *": "Bob", "Last Name *": "Expired",
        "Rifle Competency Number": "E2345678",
        "Rifle Competency Expiry": past,
    }])
    res = svc.import_workbook(db, data)
    assert res["imported"] == 1
    assert res["failed"] == 0
    g = _guard(db, "Expired")
    assert g.permitted_rifle is True              # still ticked despite expiry
    assert len(res["expired_review"]) == 1
    entry = res["expired_review"][0]
    assert entry["weapon"] == "Rifle"
    assert entry["name"] == "Bob Expired"
    assert entry["competency_number"] == "E2345678"
    assert res["error_workbook"] is None          # expiry is NOT an error


def test_bad_format_and_incomplete_pair_fail_into_error_workbook(db):
    future = date.today() + timedelta(days=365)
    data = _make_guards_book([
        {"First Name *": "Carl", "Last Name *": "BadFmt",
         "Handgun Competency Number": "XX12", "Handgun Competency Expiry": future},
        {"First Name *": "Dora", "Last Name *": "HalfPair",
         "Carbine Competency Number": "D1234567"},  # no expiry
        {"First Name *": "Eve", "Last Name *": "Good"},  # no competencies at all -> ok
    ])
    res = svc.import_workbook(db, data)
    assert res["imported"] == 1                   # only Eve
    assert res["failed"] == 2
    assert _guard(db, "BadFmt") is None
    assert _guard(db, "HalfPair") is None
    assert _guard(db, "Good") is not None

    # Error workbook present and re-importable: mirrors template + Error Details.
    assert res["error_workbook"] is not None
    wb = load_workbook(BytesIO(base64.b64decode(res["error_workbook"]["content_base64"])))
    ws = wb["Guards"]
    headers = [c.value for c in ws[1]]
    assert headers[-1] == "Error Details"
    assert headers[:-1] == GUARD_HEADERS
    reasons = [row[-1].value for row in ws.iter_rows(min_row=2)]
    assert any("invalid format" in r for r in reasons)
    assert any("without Expiry Date" in r for r in reasons)


def test_both_blank_is_not_authorised_no_error(db):
    data = _make_guards_book([{"First Name *": "Fay", "Last Name *": "Blank"}])
    res = svc.import_workbook(db, data)
    assert res["imported"] == 1
    assert res["failed"] == 0
    g = _guard(db, "Blank")
    assert not any([g.permitted_shotgun, g.permitted_carbine,
                    g.permitted_rifle, g.permitted_handgun])
