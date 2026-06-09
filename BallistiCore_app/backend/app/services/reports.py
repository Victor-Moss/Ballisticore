"""
Report generation service.
Produces in-memory Excel workbooks for SAPS-compliant audit exports.
"""
import io
from datetime import datetime, date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.register import Register
from app.models.register_history import RegisterHistory
from app.models.guard import Guard
from app.models.firearm import Firearm
from app.models.user import User


# ── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="EBF0F7")
THIN = Side(style="thin", color="C0C8D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ISSUED_FILL = PatternFill("solid", fgColor="D6EAF8")
RETURNED_FILL = PatternFill("solid", fgColor="D5F5E3")


def _apply_header(ws, headers: list[str], row: int = 1):
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws, min_width: int = 12, max_width: int = 40):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(
            min_width, min(length + 2, max_width)
        )


def _title_row(ws, title: str, col_span: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    sub = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    sub.font = Font(italic=True, size=9, color="777777")
    sub.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 15


# ── Register snapshot ────────────────────────────────────────────────────────

def generate_register_excel(db: Session) -> bytes:
    """Current register — all firearms currently issued."""
    entries = (
        db.query(Register)
        .join(Guard, Register.guard_id == Guard.id)
        .join(Firearm, Register.firearm_id == Firearm.id)
        .order_by(Guard.last_name, Guard.first_name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Current Register"
    ws.freeze_panes = "A5"

    HEADERS = [
        "Guard ID", "Guard Name", "ID Number", "PSIRA Number",
        "Firearm Make", "Firearm Model", "Serial Number", "Calibre",
        "License Number", "Issued Date", "Issued Time", "Permit Number",
    ]
    _title_row(ws, "BallistiCore — Current Firearms Register", len(HEADERS))

    # Blank spacer row
    ws.row_dimensions[3].height = 6

    _apply_header(ws, HEADERS, row=4)

    for i, entry in enumerate(entries, start=5):
        guard = entry.guard
        firearm = entry.firearm

        # Look up permit number if available
        permit_number = ""
        if entry.permit_id:
            from app.models.permit import Permit
            permit = db.query(Permit).filter(Permit.id == entry.permit_id).first()
            if permit:
                permit_number = permit.permit_number

        row_data = [
            str(guard.id)[:8] + "…",
            f"{guard.first_name} {guard.last_name}",
            guard.id_number or "",
            guard.psira_number or "",
            firearm.make or "",
            firearm.model or "",
            firearm.serial_number or "",
            firearm.calibre or "",
            firearm.license_number or "",
            entry.issued_at.strftime("%Y-%m-%d") if entry.issued_at else "",
            entry.issued_at.strftime("%H:%M") if entry.issued_at else "",
            permit_number,
        ]

        fill = ALT_FILL if i % 2 == 0 else None
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    _auto_width(ws)

    # Summary row
    summary_row = len(entries) + 5
    ws.cell(row=summary_row, column=1, value=f"Total issued: {len(entries)}").font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── History export ───────────────────────────────────────────────────────────

def generate_history_excel(
    db: Session,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    guard_id: Optional[str] = None,
    firearm_id: Optional[str] = None,
) -> bytes:
    """Filtered register history — SAPS audit trail export."""
    from datetime import datetime as dt
    q = db.query(RegisterHistory)

    if from_date:
        q = q.filter(RegisterHistory.actioned_at >= dt.combine(from_date, dt.min.time()))
    if to_date:
        q = q.filter(RegisterHistory.actioned_at <= dt.combine(to_date, dt.max.time()))
    if guard_id:
        q = q.filter(RegisterHistory.guard_id == guard_id)
    if firearm_id:
        q = q.filter(RegisterHistory.firearm_id == firearm_id)

    entries = q.order_by(RegisterHistory.actioned_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Register History"
    ws.freeze_panes = "A5"

    HEADERS = [
        "Date", "Time", "Action",
        "Guard Name", "ID Number", "PSIRA Number",
        "Firearm Make", "Firearm Model", "Serial Number", "Calibre",
        "Actioned By", "Notes",
    ]

    # Build date range label for title
    date_label = ""
    if from_date and to_date:
        date_label = f" ({from_date} to {to_date})"
    elif from_date:
        date_label = f" (from {from_date})"
    elif to_date:
        date_label = f" (to {to_date})"

    _title_row(ws, f"BallistiCore — Register History{date_label}", len(HEADERS))
    ws.row_dimensions[3].height = 6
    _apply_header(ws, HEADERS, row=4)

    # Pre-fetch guard/firearm/user lookup maps
    guard_map = {g.id: g for g in db.query(Guard).all()}
    firearm_map = {f.id: f for f in db.query(Firearm).all()}
    user_map = {u.id: u for u in db.query(User).all()}

    for i, entry in enumerate(entries, start=5):
        guard = guard_map.get(entry.guard_id)
        firearm = firearm_map.get(entry.firearm_id)
        actioned_user = user_map.get(entry.actioned_by)

        is_issued = entry.action.lower() == "issued"
        row_fill = ISSUED_FILL if is_issued else RETURNED_FILL

        row_data = [
            entry.actioned_at.strftime("%Y-%m-%d") if entry.actioned_at else "",
            entry.actioned_at.strftime("%H:%M") if entry.actioned_at else "",
            entry.action.upper(),
            f"{guard.first_name} {guard.last_name}" if guard else entry.guard_id,
            guard.id_number if guard else "",
            guard.psira_number if guard else "",
            firearm.make if firearm else "",
            firearm.model if firearm else "",
            firearm.serial_number if firearm else entry.firearm_id,
            firearm.calibre if firearm else "",
            actioned_user.username if actioned_user else str(entry.actioned_by)[:8],
            entry.notes or "",
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill = row_fill
            if col == 3:  # Action column — bold
                cell.font = Font(bold=True)

    _auto_width(ws)

    # Totals
    issued_count = sum(1 for e in entries if e.action.lower() == "issued")
    returned_count = len(entries) - issued_count
    summary_row = len(entries) + 5
    ws.cell(row=summary_row, column=1,
            value=f"Total records: {len(entries)}  |  Issued: {issued_count}  |  Returned: {returned_count}"
            ).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Guard activity report ────────────────────────────────────────────────────

def generate_guard_activity_excel(db: Session, guard_id: str) -> bytes:
    """All history for a single guard — useful for individual SAPS audits."""
    guard = db.query(Guard).filter(Guard.id == guard_id).first()
    if not guard:
        raise ValueError("Guard not found")

    entries = (
        db.query(RegisterHistory)
        .filter(RegisterHistory.guard_id == guard_id)
        .order_by(RegisterHistory.actioned_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Guard Activity"

    guard_name = f"{guard.first_name} {guard.last_name}"
    HEADERS = ["Date", "Time", "Action", "Firearm Make", "Firearm Model", "Serial Number", "Calibre", "Notes"]

    _title_row(ws, f"Guard Activity Report — {guard_name}", len(HEADERS))

    # Guard details block
    ws.row_dimensions[3].height = 6
    detail_row = 3
    ws.cell(row=detail_row, column=1, value="ID Number:").font = Font(bold=True)
    ws.cell(row=detail_row, column=2, value=guard.id_number or "—")
    ws.cell(row=detail_row, column=3, value="PSIRA:").font = Font(bold=True)
    ws.cell(row=detail_row, column=4, value=guard.psira_number or "—")

    _apply_header(ws, HEADERS, row=5)

    firearm_map = {f.id: f for f in db.query(Firearm).all()}

    for i, entry in enumerate(entries, start=6):
        firearm = firearm_map.get(entry.firearm_id)
        is_issued = entry.action.lower() == "issued"
        row_fill = ISSUED_FILL if is_issued else RETURNED_FILL

        row_data = [
            entry.actioned_at.strftime("%Y-%m-%d") if entry.actioned_at else "",
            entry.actioned_at.strftime("%H:%M") if entry.actioned_at else "",
            entry.action.upper(),
            firearm.make if firearm else "",
            firearm.model if firearm else "",
            firearm.serial_number if firearm else entry.firearm_id,
            firearm.calibre if firearm else "",
            entry.notes or "",
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill = row_fill

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
