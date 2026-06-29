"""
Bulk import from the BallistiCore Excel template.

Provides:
  - build_template()  -> bytes  : the blank .xlsx (Guards / Firearms / Users)
  - import_workbook(db, data)    : parse + validate + insert, returning a
                                   per-sheet report of imported / failed / skipped rows.

Validation is per-row: a bad row is reported with its sheet, row number and
reason, and never blocks the valid rows in the same upload.

SAPS competency: the Guards sheet carries a Competency Number + Expiry Date pair
for each weapon type (shotgun, carbine, rifle, handgun). A valid, complete pair
auto-ticks that weapon's permission on the guard. A pair whose expiry is in the
past still imports (and still ticks) but is surfaced in a "Review Expired
Competencies" list. A malformed number or a half-filled pair fails the row, which
is then excluded from the import and written to a downloadable error workbook.
"""
import base64
import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.messaging_config import get_provider
from app.models.ammunition_type import AmmunitionType
from app.schemas.firearm import FirearmCreate, FIREARM_TYPES
from app.schemas.guard import GuardCreate
from app.services import firearms as firearm_svc
from app.services import guards as guard_svc
from app.services import users as user_svc

# Rows whose first cell starts with this are treated as format examples and skipped.
EXAMPLE_PREFIX = "e.g."

# Weapon types that carry a SAPS competency pair, in the order they appear in the
# Guards sheet. Each maps to model fields saps_comp_<w> / saps_expiry_<w> /
# permitted_<w>.
WEAPON_TYPES = ["shotgun", "carbine", "rifle", "handgun"]

# A competency number is a single letter prefix followed by exactly seven digits
# (e.g. C7021766). Case-insensitive on input; normalised to uppercase on store.
_COMP_RE = re.compile(r"^[A-Za-z]\d{7}$")

# Date formats accepted when an expiry cell arrives as text rather than a real
# Excel date. Real date cells come back as datetime objects and skip this.
_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y")

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EXAMPLE_FONT = Font(italic=True, color="64748B")


def _cell_str(v) -> str:
    """Normalise a cell value to a trimmed string ('' for blanks)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _truthy(v: str) -> bool:
    return _cell_str(v).lower() in {"yes", "y", "true", "1", "x", "admin"}


class _RowInvalid(ValueError):
    """Raised when a date cell can't be parsed; message is user-facing."""


def _parse_date(v):
    """Parse an Excel cell into a date, or None for blanks. Raise _RowInvalid
    on a non-blank value that isn't a recognisable date."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise _RowInvalid("not a valid date")


# ── Creators: validate one row's values and insert. Raise on any problem. ─────
def _create_guard(db: Session, v: dict):
    # SAPS competency fields and permitted_* flags are pre-validated and
    # normalised by _validate_guard_competencies before we get here: comp numbers
    # are uppercased str|None, expiry values are date|None, permitted_* are bools.
    data = GuardCreate(
        first_name=v["first_name"], last_name=v["last_name"],
        id_number=v.get("id_number") or None,
        psira_number=v.get("psira_number") or None,
        cell_phone=v.get("cell_phone") or None,
        telegram_chat_id=v.get("telegram_chat_id") or None,
        email=v.get("email") or None,
        physical_address=v.get("physical_address") or None,
        region=v.get("region") or None,
        personnel_number=v.get("personnel_number") or None,
        saps_comp_shotgun=v.get("saps_comp_shotgun"),
        saps_expiry_shotgun=v.get("saps_expiry_shotgun"),
        saps_comp_carbine=v.get("saps_comp_carbine"),
        saps_expiry_carbine=v.get("saps_expiry_carbine"),
        saps_comp_rifle=v.get("saps_comp_rifle"),
        saps_expiry_rifle=v.get("saps_expiry_rifle"),
        saps_comp_handgun=v.get("saps_comp_handgun"),
        saps_expiry_handgun=v.get("saps_expiry_handgun"),
        permitted_shotgun=bool(v.get("permitted_shotgun")),
        permitted_carbine=bool(v.get("permitted_carbine")),
        permitted_rifle=bool(v.get("permitted_rifle")),
        permitted_handgun=bool(v.get("permitted_handgun")),
    )
    guard_svc.create(db, data)


def _resolve_ammunition_type(db: Session, name: str | None) -> str | None:
    """Map an Ammunition Type *name* from the sheet to its id. Blank -> None;
    an unknown name fails the row (the type must already exist — manage them under
    Firearms -> Ammunition Types). Matched case-insensitively against active types."""
    name = (name or "").strip()
    if not name:
        return None
    ammo = (
        db.query(AmmunitionType)
        .filter(func.lower(AmmunitionType.name) == name.lower(),
                AmmunitionType.is_active == True)  # noqa: E712
        .first()
    )
    if not ammo:
        raise ValueError(
            f"Ammunition type '{name}' not found — add it under Firearms → Ammunition Types first"
        )
    return ammo.id


def _create_firearm(db: Session, v: dict):
    # license_issue_date is pre-parsed to date|None by _validate_firearm before we
    # get here; ammunition_type arrives as a name and is resolved to its id below.
    ftype = (v.get("type") or "").lower()
    if ftype and ftype not in FIREARM_TYPES:
        raise ValueError(f"Type must be one of: {', '.join(FIREARM_TYPES)}")
    if firearm_svc.get_by_serial(db, v["serial_number"]):
        raise ValueError(f"A firearm with serial '{v['serial_number']}' already exists")
    data = FirearmCreate(
        serial_number=v["serial_number"], make=v["make"],
        model=v.get("model") or None, type=ftype or None,
        calibre=v.get("calibre") or None,
        license_number=v.get("license_number") or None,
        license_issue_date=v.get("license_issue_date"),
        description=v.get("description") or None,
        ammunition_type_id=_resolve_ammunition_type(db, v.get("ammunition_type")),
    )
    firearm_svc.create(db, data)


def _create_user(db: Session, v: dict):
    if len(v.get("password", "")) < 6:
        raise ValueError("Password must be at least 6 characters")
    if user_svc.get_by_username(db, v["username"]):
        raise ValueError(f"Username '{v['username']}' already exists")
    user_svc.create(
        db, username=v["username"], email=v.get("email") or None,
        password=v["password"], is_admin=_truthy(v.get("is_admin", "")),
        personnel_number=v.get("personnel_number") or None,
        psira_number=v.get("psira_number") or None,
        competency=v.get("competency") or None,
        phone_number=v.get("phone_number") or None,
        id_number=v.get("id_number") or None,
    )


# ── SAPS competency validation (Guards sheet) ─────────────────────────────────
_COMP_COMMENT = ("Letter + 7 digits, e.g. C7021766. Leave blank (with the Expiry) "
                 "if not authorised for this weapon.")
_DATE_COMMENT = ("Expiry date, e.g. 2026-12-31. A past date still imports but is "
                 "flagged for review. Leave blank (with the Number) if N/A.")


def _validate_guard_competencies(values: dict, raw: dict) -> tuple[list[str], list[dict]]:
    """Validate the four SAPS competency pairs on a Guards row.

    Mutates `values` in place: normalises each comp number to uppercase str|None,
    replaces each expiry with a parsed date|None, and sets permitted_<weapon> True
    for every valid, complete pair.

    `raw` holds the original (unstringified) cell values so expiry dates keep their
    native type. Returns (errors, expired) where:
      - errors  : human-readable reasons this row is invalid (empty == row is OK)
      - expired : complete & valid pairs whose expiry is in the past — these do NOT
                  make the row invalid; they're surfaced for follow-up review.
    """
    errors: list[str] = []
    expired: list[dict] = []
    today = date.today()

    for w in WEAPON_TYPES:
        label = w.capitalize()
        comp = (values.get(f"saps_comp_{w}") or "").strip().upper()

        exp_date = None
        date_invalid = False
        try:
            exp_date = _parse_date(raw.get(f"saps_expiry_{w}"))
        except _RowInvalid:
            errors.append(f"{label} Competency Expiry: invalid date")
            date_invalid = True

        has_comp = bool(comp)
        has_exp = exp_date is not None

        # Store the normalised values back (used by the creator on success).
        values[f"saps_comp_{w}"] = comp or None
        values[f"saps_expiry_{w}"] = exp_date

        # Neither side filled (and no bad date) -> not authorised for this weapon.
        if not has_comp and not has_exp and not date_invalid:
            continue

        if has_comp and not _COMP_RE.match(comp):
            errors.append(f"{label} Competency Number: invalid format")
        if has_comp and not has_exp and not date_invalid:
            errors.append(f"{label}: Competency Number present without Expiry Date")
        if has_exp and not has_comp:
            errors.append(f"{label}: Expiry Date present without Competency Number")

        # Auto-tick only a genuinely complete, well-formed pair.
        if has_comp and _COMP_RE.match(comp) and has_exp:
            values[f"permitted_{w}"] = True
            if exp_date < today:
                expired.append({
                    "weapon": label,
                    "competency_number": comp,
                    "expiry_date": exp_date.isoformat(),
                })

    return errors, expired


def _validate_firearm(values: dict, raw: dict) -> tuple[list[str], list[dict]]:
    """Parse the Firearms sheet's Licence Issue Date from its native cell value,
    storing a date|None back into `values`. Returns (errors, []) — firearms have
    no expired-review concept, so the second list is always empty."""
    errors: list[str] = []
    try:
        values["license_issue_date"] = _parse_date(raw.get("license_issue_date"))
    except _RowInvalid:
        errors.append("Licence Issue Date: invalid date")
    return errors, []


# ── Sheet definitions ─────────────────────────────────────────────────────────
# Each column: (header, field, required, example, comment)
#
# The Guards sheet's delivery-contact column depends on the configured messaging
# provider (Settings → Messaging): a Telegram Chat ID column, a WhatsApp/Cell
# Phone column, or no contact column at all (provider "none"). The sheet set is
# therefore built per-request via build_sheets() rather than fixed at import time.

_TELEGRAM_COMMENT = ("The guard's Telegram Chat ID. The guard must first send /start "
                     "to your company's Telegram bot, which replies with their Chat ID.")


def _guard_contact_column(provider: str):
    """The provider-appropriate delivery-contact column for the Guards sheet,
    or None when the provider is 'none' (no delivery field needed)."""
    if provider == "telegram":
        return ("Telegram Chat ID", "telegram_chat_id", False, "123456789", _TELEGRAM_COMMENT)
    if provider == "whatsapp":
        return ("Cell Phone", "cell_phone", False, "0821234567", None)
    return None


def _guard_columns(provider: str):
    cols = [
        ("First Name *",       "first_name",       True,  "e.g. John",                None),
        ("Last Name *",        "last_name",        True,  "Smith",                    None),
        ("ID Number",          "id_number",        False, "8001015009087",            None),
        ("PSIRA Number",       "psira_number",     False, "PS1234567",                "Any value (free text). Leave blank if unknown."),
    ]
    contact = _guard_contact_column(provider)
    if contact:
        cols.append(contact)
    cols += [
        ("Email",              "email",            False, "john.smith@example.com",   None),
        ("Physical Address",   "physical_address", False, "12 Main Rd, Springs",      None),
        ("Region",             "region",           False, "Gauteng",                  "Any value (free text). Leave blank if unknown."),
        ("Personnel Number",   "personnel_number", False, "EMP-001",                  None),
        # SAPS competency: a number + expiry pair per weapon type. Leave both
        # blank if the guard isn't authorised for that weapon. A complete,
        # valid pair auto-ticks that weapon's permission on import.
        ("Shotgun Competency Number", "saps_comp_shotgun",   False, "C7021766",   _COMP_COMMENT),
        ("Shotgun Competency Expiry", "saps_expiry_shotgun", False, "2026-12-31", _DATE_COMMENT),
        ("Carbine Competency Number", "saps_comp_carbine",   False, "D1234567",   _COMP_COMMENT),
        ("Carbine Competency Expiry", "saps_expiry_carbine", False, "2026-12-31", _DATE_COMMENT),
        ("Rifle Competency Number",   "saps_comp_rifle",     False, "E2345678",   _COMP_COMMENT),
        ("Rifle Competency Expiry",   "saps_expiry_rifle",   False, "2026-12-31", _DATE_COMMENT),
        ("Handgun Competency Number", "saps_comp_handgun",   False, "F3456789",   _COMP_COMMENT),
        ("Handgun Competency Expiry", "saps_expiry_handgun", False, "2026-12-31", _DATE_COMMENT),
    ]
    return cols


def build_sheets(provider: str | None = None) -> list[dict]:
    """The import/export sheet definitions. The Guards contact column reflects the
    active messaging provider (defaults to the currently configured one)."""
    if provider is None:
        provider = get_provider()
    return [
        {
            "name": "Guards",
            "creator": _create_guard,
            "validator": _validate_guard_competencies,
            "columns": _guard_columns(provider),
        },
        {
            "name": "Firearms",
            "creator": _create_firearm,
            "validator": _validate_firearm,
            "columns": [
                ("Serial Number *",    "serial_number",      True,  "e.g. GLK-100234", None),
                ("Make *",             "make",               True,  "Glock",           None),
                ("Model",              "model",              False, "17",              None),
                ("Type",               "type",               False, "handgun",         "One of: carbine, handgun, rifle, shotgun (or leave blank)"),
                ("Calibre",            "calibre",            False, "9mm",             None),
                ("Licence Number",     "license_number",     False, "LIC-2024-001",    None),
                ("Licence Issue Date", "license_issue_date", False, "2024-01-15",      "Issue date, e.g. 2024-01-15. Leave blank if unknown."),
                ("Ammunition Type",    "ammunition_type",    False, "9mm FMJ",         "Must match an existing Ammunition Type name (Firearms → Ammunition Types). Leave blank if none."),
                ("Description",        "description",        False, "Duty pistol",     None),
            ],
        },
        {
            "name": "Users",
            "creator": _create_user,
            "columns": [
                ("Username *",         "username",         True,  "e.g. joperator",          None),
                ("Password *",         "password",         True,  "ChangeMe!23",             "At least 6 characters"),
                ("Email",              "email",            False, "joperator@example.com",   None),
                ("Is Admin",           "is_admin",         False, "no",                      "yes = full administrator, no = operator"),
                ("Personnel Number",   "personnel_number", False, "EMP-100",                 None),
                ("PSIRA Number",       "psira_number",     False, "PS9988776",               None),
                ("Competency",         "competency",       False, "Grade C",                 None),
                ("Phone Number",       "phone_number",     False, "0837654321",              None),
                ("ID Number",          "id_number",        False, "9002025009088",           None),
            ],
        },
    ]


# ── Template generation ───────────────────────────────────────────────────────
def build_template() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    for sheet in build_sheets():
        ws = wb.create_sheet(title=sheet["name"])
        cols = sheet["columns"]
        # Header row
        for c, (header, _field, _req, _ex, comment) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if comment:
                cell.comment = Comment(comment, "BallistiCore")
            ws.column_dimensions[get_column_letter(c)].width = max(16, len(header) + 4)
        # Example row (ignored on import — first cell starts with "e.g.")
        for c, (_h, _f, _r, example, _cm) in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=c, value=example)
            cell.font = _EXAMPLE_FONT
        ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import ────────────────────────────────────────────────────────────────────
def _friendly_error(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        parts = []
        for e in exc.errors():
            loc = ".".join(str(x) for x in e.get("loc", ())) or "value"
            parts.append(f"{loc}: {e.get('msg', 'invalid')}")
        return "; ".join(parts)
    if isinstance(exc, IntegrityError):
        return "Duplicate or invalid value (rejected by a database constraint)"
    return str(exc) or exc.__class__.__name__


def _row_cells(row, header_to_field, cols) -> list:
    """Original cell values for one row, in the sheet's template column order.

    Used to rebuild a failed row in the error workbook so the user can fix just
    that cell and re-import. Keeps native types (dates stay dates)."""
    out = []
    for _h, field, *_ in cols:
        idx = header_to_field.get(field)
        out.append(row[idx] if idx is not None and idx < len(row) else None)
    return out


def _build_error_workbook(error_sheets: dict) -> bytes:
    """Build an .xlsx of every failed row, mirroring the template's columns plus
    a trailing 'Error Details' column explaining why each row failed."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, info in error_sheets.items():
        ws = wb.create_sheet(title=name)
        headers = [h for (h, *_ ) in info["columns"]] + ["Error Details"]
        for c, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = max(16, len(header) + 4)
        for r, (cells, message) in enumerate(info["rows"], start=2):
            for c, val in enumerate(cells, start=1):
                ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=len(cells) + 1, value=message)
        ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_workbook(db: Session, data: bytes) -> dict:
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception:
        raise ValueError("That file is not a valid .xlsx workbook.")

    results = []
    total_imported = total_failed = 0
    expired_review: list[dict] = []
    error_sheets: dict = {}  # sheet name -> {"columns": cols, "rows": [(cells, msg)]}

    for sheet in build_sheets():
        name = sheet["name"]
        cols = sheet["columns"]
        validator = sheet.get("validator")
        imported = 0
        skipped = 0
        errors = []

        if name not in wb.sheetnames:
            results.append({"sheet": name, "imported": 0, "failed": 0,
                            "skipped": 0, "errors": [],
                            "note": "Sheet not found in the uploaded file."})
            continue

        ws = wb[name]
        # Map by header text so column order in the file doesn't matter.
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_to_field = {}
        for header, field, *_ in cols:
            for idx, hc in enumerate(header_cells):
                if _cell_str(hc).lower() == header.lower():
                    header_to_field[field] = idx
                    break

        def _record_error(excel_row, row, message):
            """Add a row to both the JSON report and the error workbook."""
            errors.append({"row": excel_row, "message": message})
            error_sheets.setdefault(name, {"columns": cols, "rows": []})
            error_sheets[name]["rows"].append((_row_cells(row, header_to_field, cols), message))

        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = {field: _cell_str(row[idx]) if idx < len(row) else ""
                      for field, idx in header_to_field.items()}
            if not any(values.values()):
                continue  # fully blank row
            first = next(iter(values.values()), "")
            if first.lower().startswith(EXAMPLE_PREFIX):
                skipped += 1
                continue
            missing = [h for (h, f, req, *_ ) in cols if req and not values.get(f)]
            if missing:
                _record_error(excel_row, row, f"Missing required: {', '.join(missing)}")
                continue

            # Per-sheet structured validation (SAPS competency pairs on Guards).
            # The validator normalises `values` and reports field-level errors;
            # it also returns expired-but-valid pairs to surface after import.
            row_expired = []
            if validator:
                raw = {field: (row[idx] if idx < len(row) else None)
                       for field, idx in header_to_field.items()}
                field_errors, row_expired = validator(values, raw)
                if field_errors:
                    _record_error(excel_row, row, "; ".join(field_errors))
                    continue

            try:
                sheet["creator"](db, values)
                imported += 1
                for e in row_expired:
                    expired_review.append({
                        "sheet": name, "row": excel_row,
                        "name": f"{values.get('first_name', '')} {values.get('last_name', '')}".strip(),
                        **e,
                    })
            except Exception as exc:  # noqa: BLE001 — per-row isolation is intentional
                db.rollback()
                _record_error(excel_row, row, _friendly_error(exc))

        total_imported += imported
        total_failed += len(errors)
        results.append({"sheet": name, "imported": imported,
                        "failed": len(errors), "skipped": skipped, "errors": errors})

    wb.close()

    response = {
        "imported": total_imported,
        "failed": total_failed,
        "expired_review": expired_review,
        "sheets": results,
        "error_workbook": None,
    }
    if error_sheets:
        wb_bytes = _build_error_workbook(error_sheets)
        response["error_workbook"] = {
            "filename": "BallistiCore_Import_Errors.xlsx",
            "content_base64": base64.b64encode(wb_bytes).decode("ascii"),
        }
    return response
