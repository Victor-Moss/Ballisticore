"""
Full data export for backup, client handoff, compliance and re-import.

Produces, in one action, three in-memory artefacts covering every entity in the
system:

  - build_excel(datasets)    -> bytes  : one .xlsx, one sheet per entity
  - build_csv_zip(datasets)  -> bytes  : one .zip of one CSV per entity
  - build_pdf_summary(...)   -> bytes  : a human-readable compliance summary

`build_full_export(db, generated_by)` bundles all three into a single ZIP and is
what the router serves.

The Guards sheet/CSV columns are derived from the import template's SHEETS config
(see app.services.imports) rather than re-listed here, so the import template and
the export stay in lock-step as guard fields are added. Every other entity is a
straight one-sheet-per-table dump of its current columns, which automatically
picks up new columns without changes here.
"""
import csv
import io
import logging
import zipfile
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.orm import Session

from app.models.ammunition_type import AmmunitionType
from app.models.firearm import Firearm
from app.models.guard import Guard, GuardCITRoute
from app.models.location import Location
from app.models.permission import GuardFirearmPermission
from app.models.permit import Permit
from app.models.register import Register
from app.models.register_history import RegisterHistory
from app.models.user import User
from app.services.imports import build_sheets, WEAPON_TYPES

logger = logging.getLogger("ballisticore.export")

# Columns never written to any export — password hashes and OTP secrets.
_SECRET_COLUMNS = {"hashed_password", "reset_otp_hash"}

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Entities exported as a straight dump of their table columns. Guards is handled
# separately so its columns track the import template. (sheet name, model).
_TABLE_ENTITIES = [
    ("Users", User),
    ("Firearms", Firearm),
    ("Locations", Location),
    ("Ammunition Types", AmmunitionType),
    ("Guard Firearm Permissions", GuardFirearmPermission),
    ("Guard CIT Routes", GuardCITRoute),
    ("Permits", Permit),
    ("Register", Register),
    ("Register History", RegisterHistory),
]


def _cell(v):
    """Serialise a value to a flat, display/CSV-friendly form."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return str(v)


def _csv_filename(sheet_name: str) -> str:
    return sheet_name.lower().replace(" ", "_") + ".csv"


# ── Human-readable companion columns for foreign keys ─────────────────────────
# Raw FK columns export as opaque UUIDs (fine for round-trip), so each FK column
# gets a readable companion column immediately to its right. To avoid a per-row DB
# lookup, every related table's id→label mapping is fetched once per export run
# (see _build_fk_maps) and looked up in memory while building each sheet/CSV.

def _firearm_label(f) -> str:
    """Best human identifier for a firearm: serial + make/model + type."""
    descr = " ".join(p for p in (f.make, f.model) if p)
    extra = ", ".join(p for p in (descr, f.type) if p)
    return f"{f.serial_number} ({extra})" if extra else (f.serial_number or "")


def _build_fk_maps(db: Session) -> dict:
    """id → readable-label maps for every table referenced by a foreign key,
    fetched once per export. Keyed by the referenced table name."""
    return {
        "guards": {g.id: f"{g.first_name} {g.last_name}".strip() for g in db.query(Guard).all()},
        "users": {u.id: u.username for u in db.query(User).all()},
        "firearms": {f.id: _firearm_label(f) for f in db.query(Firearm).all()},
        "ammunition_types": {a.id: a.name for a in db.query(AmmunitionType).all()},
        "permits": {p.id: p.permit_number for p in db.query(Permit).all()},
        "locations": {l.id: l.name for l in db.query(Location).all()},
    }

# What the companion column is named, by referenced table.
_COMPANION_SUFFIX = {
    "guards": "name", "users": "name", "locations": "name",
    "ammunition_types": "name", "permits": "number", "firearms": "description",
}


def _companion_header(fk_column: str, target_table: str) -> str:
    base = fk_column[:-3] if fk_column.endswith("_id") else fk_column
    return f"{base}_{_COMPANION_SUFFIX[target_table]}"


def _guards_dataset(db: Session) -> dict:
    """Guards columns derived from the import template (header, field) pairs, so
    this export round-trips back through the import."""
    guard_sheet = next(s for s in build_sheets() if s["name"] == "Guards")
    cols = [(header, field) for (header, field, *_ ) in guard_sheet["columns"]]
    headers = [h for h, _ in cols]
    rows = [
        [_cell(getattr(g, field, None)) for _, field in cols]
        for g in db.query(Guard).order_by(Guard.last_name, Guard.first_name).all()
    ]
    return {"name": "Guards", "filename": _csv_filename("Guards"), "headers": headers, "rows": rows}


def _table_dataset(db: Session, name: str, model, fk_maps: dict) -> dict:
    """Dump a table's columns; for each foreign key, add a readable companion
    column right after the raw ID column (resolved against the in-memory maps).
    The original ID column is always kept intact for backup/round-trip."""
    # Plan the output columns: each is ("plain", col_name) or ("fk", col_name, mapping).
    plan: list[tuple] = []
    headers: list[str] = []
    for c in model.__table__.columns:
        if c.name in _SECRET_COLUMNS:
            continue
        plan.append(("plain", c.name))
        headers.append(c.name)
        if c.foreign_keys:
            target = next(iter(c.foreign_keys)).column.table.name
            mapping = fk_maps.get(target)
            if mapping is not None:
                plan.append(("fk", c.name, mapping))
                headers.append(_companion_header(c.name, target))

    rows = []
    for obj in db.query(model).all():
        row = []
        for entry in plan:
            if entry[0] == "plain":
                row.append(_cell(getattr(obj, entry[1])))
            else:  # "fk" companion
                fk_value = getattr(obj, entry[1])
                row.append(entry[2].get(fk_value, "") if fk_value else "")
        rows.append(row)
    return {"name": name, "filename": _csv_filename(name), "headers": headers, "rows": rows}


def collect_datasets(db: Session) -> list[dict]:
    """One {name, filename, headers, rows} per entity — the single source of data
    that both the Excel and CSV exports render."""
    fk_maps = _build_fk_maps(db)  # fetched once, reused across every sheet
    datasets = [_guards_dataset(db)]
    for name, model in _TABLE_ENTITIES:
        datasets.append(_table_dataset(db, name, model, fk_maps))
    return datasets


# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(datasets: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for ds in datasets:
        ws = wb.create_sheet(title=ds["name"][:31])  # Excel caps sheet names at 31 chars
        ws.append(ds["headers"])
        for c, header in enumerate(ds["headers"], start=1):
            cell = ws.cell(row=1, column=c)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(header) + 4, 40))
        for row in ds["rows"]:
            ws.append(row)
        ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV bundle ──────────────────────────────────────────────────────────────────
def build_csv_zip(datasets: list[dict]) -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for ds in datasets:
            s = io.StringIO()
            writer = csv.writer(s)
            writer.writerow(ds["headers"])
            writer.writerows(ds["rows"])
            z.writestr(ds["filename"], s.getvalue())
    return buf.getvalue()


# ── PDF compliance summary ──────────────────────────────────────────────────────
def _summary_stats(db: Session) -> dict:
    guards = db.query(Guard).all()
    today = date.today()

    permission_breakdown = {
        w.capitalize(): sum(1 for g in guards if getattr(g, f"permitted_{w}"))
        for w in WEAPON_TYPES
    }

    expired = []
    for g in guards:
        for w in WEAPON_TYPES:
            number = getattr(g, f"saps_comp_{w}")
            expiry = getattr(g, f"saps_expiry_{w}")
            if number and expiry and expiry < today:
                expired.append({
                    "name": f"{g.first_name} {g.last_name}".strip(),
                    "weapon": w.capitalize(),
                    "number": number,
                    "expiry": expiry.strftime("%Y-%m-%d"),
                })
    expired.sort(key=lambda e: (e["expiry"], e["name"]))

    with_psira = sum(1 for g in guards if (g.psira_number or "").strip())

    users = db.query(User).all()
    return {
        "total_guards": len(guards),
        "active_guards": sum(1 for g in guards if g.is_active),
        "permission_breakdown": permission_breakdown,
        "expired": expired,
        "psira_with": with_psira,
        "psira_without": len(guards) - with_psira,
        "total_users": len(users),
        "active_users": sum(1 for u in users if u.is_active),
        "admin_users": sum(1 for u in users if u.is_admin or u.perm_system_admin),
    }


def build_pdf_summary(db: Session, generated_by: str, generated_at: datetime) -> bytes:
    stats = _summary_stats(db)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm,
        title="BallistiCore Compliance Summary",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#1E3A5F"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#1D4ED8"),
                        spaceBefore=10, spaceAfter=4)
    meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#666666"))
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, alignment=TA_LEFT)

    def kv_table(rows):
        t = Table(rows, colWidths=[90 * mm, 70 * mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#333333")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ]))
        return t

    elements = [
        Paragraph("BallistiCore — Compliance Summary", h1),
        Paragraph(f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M')} &nbsp;|&nbsp; By: {generated_by}", meta),
        Spacer(1, 8),
    ]

    elements.append(Paragraph("Guards", h2))
    elements.append(kv_table([
        ["Total guards", str(stats["total_guards"])],
        ["Active guards", str(stats["active_guards"])],
    ]))

    elements.append(Paragraph("Weapon permission breakdown", h2))
    elements.append(kv_table([[w, str(n)] for w, n in stats["permission_breakdown"].items()]))

    elements.append(Paragraph("PSIRA registration coverage", h2))
    elements.append(kv_table([
        ["Guards with a PSIRA number", str(stats["psira_with"])],
        ["Guards without a PSIRA number", str(stats["psira_without"])],
    ]))

    elements.append(Paragraph("User accounts", h2))
    elements.append(kv_table([
        ["Total users", str(stats["total_users"])],
        ["Active users", str(stats["active_users"])],
        ["Admin-level users", str(stats["admin_users"])],
    ]))

    elements.append(Paragraph(f"Expired competencies ({len(stats['expired'])})", h2))
    if stats["expired"]:
        rows = [["Guard", "Weapon", "Competency No.", "Expired"]]
        rows += [[e["name"], e["weapon"], e["number"], e["expiry"]] for e in stats["expired"]]
        t = Table(rows, colWidths=[60 * mm, 28 * mm, 42 * mm, 30 * mm], hAlign="LEFT", repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B45309")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FEF3C7")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No expired competencies. ✓", body))

    doc.build(elements)
    return buf.getvalue()


# ── Full bundle ─────────────────────────────────────────────────────────────────
def build_full_export(db: Session, generated_by: str) -> tuple[bytes, str]:
    """Build the Excel workbook, CSV bundle and PDF summary, package them into a
    single ZIP, and log the export for audit. Returns (zip_bytes, filename)."""
    generated_at = datetime.now()
    stamp = generated_at.strftime("%Y%m%d_%H%M")

    datasets = collect_datasets(db)
    excel_bytes = build_excel(datasets)
    csv_zip_bytes = build_csv_zip(datasets)
    pdf_bytes = build_pdf_summary(db, generated_by=generated_by, generated_at=generated_at)

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"BallistiCore_Data_{stamp}.xlsx", excel_bytes)
        z.writestr(f"BallistiCore_CSV_{stamp}.zip", csv_zip_bytes)
        z.writestr(f"BallistiCore_Compliance_Summary_{stamp}.pdf", pdf_bytes)

    # Audit trail: who exported the full dataset, and when.
    logger.info("Full data export generated by user=%r at %s (entities=%d)",
                generated_by, generated_at.isoformat(timespec="seconds"), len(datasets))

    return buf.getvalue(), f"BallistiCore_Export_{stamp}.zip"
